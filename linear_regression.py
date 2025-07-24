"""
This script performs simple linear regression analysis
in this version the progam is able to work in multidimensional space.
It includes hypothesis testing and plotting of the data if possible.
"""
import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from scipy import stats

def data_loader(file) :
    # Reads the excel file and returns the names of the variables and their reletive shuffeled data.
    # Load data from Excel
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    x_values = []
    y_values = []

    for row in ws.iter_rows(min_row=2, values_only=True) :
        try :
            x = [float(i) for i in row]
            x_values.append(x)
        except (TypeError, ValueError):
            continue # Skip rows with invalid data

    np.random.seed(42)
    np.random.shuffle(x_values)
    for i in x_values :
        y_values.append(i[-1])
    x_values = [i[:-1] for i in x_values]

    wb.close()
    return headers, np.array(x_values), np.array(y_values)

class Linear_regression:
    def __init__(self):
        self.w = None   # Parameter vector
        self.n = None   # Number of samples
        self.x = None   # Input data
        self.y = None   # Output data

    def build(self, x_values, y_values):
        """
        Fits the linear regression model to the data using the closed-form solution.
        """
        self.n = len(x_values)

        # Add bias term (intercept) column of ones
        self.x = np.c_[np.ones((self.n, 1)), x_values]
        self.y = y_values

        # Closed-form solution: w = (X^T X)^-1 X^T y
        self.w = np.linalg.inv(self.x.T @ self.x) @ self.x.T @ self.y

    def parameters(self):
        """Returns the learned weight vector."""
        return self.w

    def predict(self, xin):
        """
        Predicts output for new input data.
        """
        xin = np.c_[np.ones((xin.shape[0], 1)), xin]
        return xin @ self.w

    def equation(self):
        """
        Returns the equation of the model as a string.
        """
        eq_str = f'{self.w[0]:.4f}'
        for i, coef in enumerate(self.w[1:], start=1):
            if coef != 0:
                eq_str += f' + {coef:.4f}x_{i}'
        return eq_str

    def mse(self):
        """
        Returns the Mean Squared Error with degrees of freedom correction (unbiased estimate).
        """
        error = (self.x @ self.w) - self.y
        return np.sum(error ** 2) / (self.n - len(self.w))

    def r_squered(self):
        """
        Returns the R² score of the model. which shows how well the points relate to the model.
        """
        ss_res = np.sum((self.y - self.x @ self.w) ** 2)
        ss_tot = np.sum((self.y - np.mean(self.y)) ** 2)
        return 1 - (ss_res / ss_tot)

    def hypothesis_test(self, i):
        """
        Returns whether the i-th parameter is statistically significant
        using a two-tailed t-test (H0: parameter = 0).
        """
        # Inverse of X^T X for standard error computation
        xtx_inv = np.linalg.inv(self.x.T @ self.x)
        # Standard error of each coefficient
        std_w = np.sqrt(self.mse() * np.diag(xtx_inv))
        # Critical t-value (two-tailed, alpha=0.05)
        t_crit = stats.t.ppf(1 - 0.025, df=self.n - len(self.w))
        # Test statistic for i-th parameter
        t_stat = abs(self.w[i] / std_w[i])
        return t_stat > t_crit

    def plot(self, headers):
        """
        Plots the fitted regression model and original data.
        Automatically selects 2D or 3D plot based on feature count.
        """
        if self.w is None:
            raise ValueError("Model not built. Call .build(x, y) before plotting.")

        x_values = self.x[:, 1:]  # Remove bias term
        y_values = self.y
        w = self.w
        r2 = self.r_squered()
        mse = self.mse()
        rejects = [self.hypothesis_test(i) for i in range(len(w))]

        n_features = x_values.shape[1]

        # --- 2D Plot ---
        if n_features == 1:
            x_min, x_max = x_values.min(), x_values.max()
            x_range = np.linspace(x_min, x_max, 100).reshape(-1, 1)
            x_range_bias = np.c_[np.ones((100, 1)), x_range]
            y_pred = x_range_bias @ w
            plt.figure(figsize=(10, 6))
            plt.scatter(x_values, y_values, color='blue', label='Data Points')
            plt.plot(x_range, y_pred, color='red', label=f"y = {w[1]:.3f}x + {w[0]:.3f}")
            info_text = (
                f"$R^2$ = {r2:.3f}\n"
                f"MSE = {mse:.3f}\n"
                f"H₀: w₀ = 0 → {'Rejected' if rejects[0] else 'Not rejected'}\n"
                f"H₀: w₁ = 0 → {'Rejected' if rejects[1] else 'Not rejected'}")
            plt.text(0.05, 0.95, info_text,
                    transform=plt.gca().transAxes,
                    fontsize=10,
                    verticalalignment='top',
                    bbox=dict(boxstyle='round', facecolor='white', alpha=0.7))

            plt.xlabel(headers[0])
            plt.ylabel(headers[-1])
            plt.title("2D Linear Regression")
            plt.legend()
            plt.grid(True)
            plt.show()

        # --- 3D Plot ---
        elif n_features == 2:
            fig = plt.figure(figsize=(10, 7))
            ax = fig.add_subplot(111, projection='3d')
            x1 = x_values[:, 0]
            x2 = x_values[:, 1]
            ax.scatter(x1, x2, y_values, color='blue', label='Data Points')
            x1_range = np.linspace(x1.min(), x1.max(), 30)
            x2_range = np.linspace(x2.min(), x2.max(), 30)
            x1_grid, x2_grid = np.meshgrid(x1_range, x2_range)
            x1x2_flat = np.c_[np.ones(x1_grid.size), x1_grid.ravel(), x2_grid.ravel()]
            y_pred_flat = x1x2_flat @ w
            y_grid = y_pred_flat.reshape(x1_grid.shape)
            ax.plot_surface(x1_grid, x2_grid, y_grid, color='red', alpha=0.5)
            ax.set_xlabel(headers[0])
            ax.set_ylabel(headers[1])
            ax.set_zlabel(headers[-1])
            ax.set_title("3D Linear Regression")

            info_text = (
                f"$R^2$ = {r2:.3f}\n"
                f"MSE = {mse:.3f}\n"
                f"H₀: w₀ = 0 → {'Rejected' if rejects[0] else 'Not rejected'}\n"
                f"H₀: w₁ = 0 → {'Rejected' if rejects[1] else 'Not rejected'}\n"
                f"H₀: w₂ = 0 → {'Rejected' if rejects[2] else 'Not rejected'}")
            ax.text2D(0.05, 0.95, info_text, transform=ax.transAxes, fontsize=10,
                    bbox=dict(boxstyle='round', facecolor='white', alpha=0.7))
            plt.tight_layout()
            plt.show()

        # --- Too many features ---
        else:
            print("❌ Cannot plot more than 2 input features.")
            print("Try reducing features using PCA or select only two.")

if __name__ == '__main__':
    headers, x, y = data_loader('points.xlsx')
    model = Linear_regression()
    model.build(x, y)
    model.plot(headers)
